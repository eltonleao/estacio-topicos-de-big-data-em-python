import json
import requests
import pandas as pd
import openpyxl

# Carregar a planilha com os CEPs, SKUs e Sellers sem cabeçalhos
input_file = "input_ceps_skus.xlsx"
ceps = pd.read_excel(input_file, sheet_name="CEPs", header=None)  # Sem cabeçalho
skus = pd.read_excel(input_file, sheet_name="SKUs", header=None)  # Sem cabeçalho
sellers = pd.read_excel(input_file, sheet_name="SELLERs", header=None)  # Sem cabeçalho

# Verifique se as colunas têm dados corretos
if ceps.empty or skus.empty or sellers.empty:
    print("Verifique o conteúdo das planilhas de CEPs, SKUs ou Sellers. Uma delas está vazia.")
    exit()

# Renomear as colunas manualmente
ceps.columns = ['CEP']
skus.columns = ['SKU']
sellers.columns = ['SELLERs']  # Certifique-se de que este nome seja usado consistentemente

# Criação da planilha de saída
excel_book = openpyxl.Workbook()
sheet = excel_book.active
sheet.title = "Dados Extração VTEX"

# Definindo cabeçalhos na planilha de saída
sheet["A1"] = "CEP"
sheet["B1"] = "SKU"
sheet["C1"] = "SELLER ID"
sheet["D1"] = "TRANSPORTADORA"
sheet["E1"] = "TEMPO"
sheet["F1"] = "CUSTO"
sheet["G1"] = "PREÇO ORIGINAL"
sheet["H1"] = "PREÇO ATUAL"
sheet["I1"] = "DISPONIBILIDADE"
sheet["J1"] = "IMPOSTO"
sheet["K1"] = "OPÇÕES DE PAGAMENTO"
sheet["L1"] = "CEP ATENDIDO"

current_row = 2

# URL da API VTEX
vtex_api_url = "https://cassol.vtexcommercestable.com.br/api/checkout/pub/orderForms/simulation"

# Cabeçalhos para a requisição
headers = {
    'Content-Type': 'application/json',
    'Accept': 'application/json',
}

def fetch_shipping_info(sku, cep, seller_id):
    payload = {
        "items": [
            {
                "id": str(sku),  # Converter SKU para string
                "quantity": 1,
                "seller": str(seller_id)  # Adicionando o Seller ID
            }
        ],
        "country": "BRA",
        "postalCode": str(cep)  # Converter CEP para string
    }

    # Faz a requisição POST à API da VTEX
    response = requests.post(vtex_api_url, json=payload, headers=headers)
    
    # Verifica se a requisição foi bem-sucedida
    if response.status_code == 200:
        data = response.json()
        json_filename = f"response_cep_{cep}_sku_{sku}_seller_{seller_id}.json"
        with open(json_filename, 'w') as f:
            json.dump(data, f, indent=4)
        print(f"Resposta da API salva em: {json_filename}")

        resultados = []
        
        if not data.get('items'):
            return [(seller_id, 'Não informado', 'Não informado', 'Não informado', 0, 0, 'Não informado', 0, 'Não informado', 'Não informado')]

        # Preço original e preço atual
        item = data['items'][0]
        preco_original = item.get('listPrice', 0) / 100 if item.get('listPrice') else 0
        preco_atual = item.get('sellingPrice', 0) / 100 if item.get('sellingPrice') else 0
        disponibilidade = item.get('availability', 'Não informado')
        imposto = item.get('tax', 0) / 100 if item.get('tax') else 0

        # Verifica se o CEP é atendido
        cep_atendido = 'Não atendido' if disponibilidade == 'cannotBeDelivered' else 'Atendido'

        # Opções de pagamento
        pagamento_info = data.get('paymentData', {}).get('installmentOptions', [])
        opcoes_pagamento = []
        for pagamento in pagamento_info:
            sistema_pagamento = pagamento.get('paymentName', 'Não informado')
            opcoes_pagamento.append(sistema_pagamento)
        
        opcoes_pagamento_str = ', '.join(opcoes_pagamento)

        # Extrai as informações de logística (transportadora, tempo e preço)
        logistics_info = data.get('logisticsInfo', [])
        if logistics_info and logistics_info[0].get('slas', []):
            for sla in logistics_info[0].get('slas', []):
                delivery_channel = sla.get('deliveryChannel', '')
                
                # Verifica se o SLA é para "delivery" (entrega)
                if delivery_channel == 'delivery':
                    transportadora = sla.get('name', 'Entrega')
                    tempo = sla.get('shippingEstimate', 'Não informado')
                    preco = sla.get('price', 0) / 100  # Convertendo centavos para reais
                    preco_formatado = f"R$ {preco:.2f}" if preco > 0 else 'Grátis'

                    resultados.append((seller_id, transportadora, tempo, preco_formatado, preco_original, preco_atual, disponibilidade, imposto, opcoes_pagamento_str, cep_atendido))
                elif delivery_channel == 'pickup-in-point':
                    transportadora = sla.get('name', 'Retirada em loja')
                    tempo = sla.get('shippingEstimate', 'Não informado')
                    preco_formatado = 'Grátis'

                    resultados.append((seller_id, transportadora, tempo, preco_formatado, preco_original, preco_atual, disponibilidade, imposto, opcoes_pagamento_str, cep_atendido))
        else:
            resultados.append((seller_id, 'Não informado', 'Não informado', 'Não informado', preco_original, preco_atual, disponibilidade, imposto, opcoes_pagamento_str, cep_atendido))

        return resultados
    
    return [(seller_id, 'Não informado', 'Não informado', 'Não informado', 0, 0, 'Não informado', 0, 'Não informado', 'Não informado')]

# Laço para iterar sobre os CEPs, SKUs e Sellers
for i, row_cep in ceps.iterrows():
    num_cep = row_cep['CEP']
    for j, row_sku in skus.iterrows():
        sku = row_sku['SKU']
        for k, row_seller in sellers.iterrows():
            seller_id = row_seller['SELLERs']

            # Busca as informações de envio
            resultados = fetch_shipping_info(sku, num_cep, seller_id)

            # Insere as informações na planilha
            for resultado in resultados:
                seller_id, transportadora, tempo, preco, preco_original, preco_atual, disponibilidade, imposto, opcoes_pagamento, cep_atendido = resultado
                sheet.cell(row=current_row, column=1, value=num_cep)
                sheet.cell(row=current_row, column=2, value=sku)
                sheet.cell(row=current_row, column=3, value=seller_id or 'Não informado')
                sheet.cell(row=current_row, column=4, value=transportadora or 'Não encontrado')
                sheet.cell(row=current_row, column=5, value=tempo or 'Não encontrado')
                sheet.cell(row=current_row, column=6, value=preco or 'Não encontrado')
                sheet.cell(row=current_row, column=7, value=preco_original or 'Não encontrado')
                sheet.cell(row=current_row, column=8, value=preco_atual or 'Não encontrado')
                sheet.cell(row=current_row, column=9, value=disponibilidade or 'Não encontrado')
                sheet.cell(row=current_row, column=10, value=imposto or 'Não encontrado')
                sheet.cell(row=current_row, column=11, value=opcoes_pagamento or 'Não encontrado')
                sheet.cell(row=current_row, column=12, value=cep_atendido or 'Não informado')

                current_row += 1

# Salva o resultado final na planilha
excel_book.save("dados_transportadoras_vtex_completos.xlsx")

print("Processo concluído!")
