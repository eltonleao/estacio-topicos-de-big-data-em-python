from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from selenium.webdriver.firefox.options import Options
import pandas as pd
import openpyxl
from time import sleep
from tqdm import tqdm


driver = webdriver.Chrome()

links = pd.read_excel(r"C:\Users\victor.hanioka\Desktop\Novo(a) Planilha do Microsoft Excel.xlsx")
ceps = pd.read_excel(r"C:\Users\victor.hanioka\Desktop\Novo(a) Planilha do Microsoft Excel.xlsx",sheet_name="ceps")

excel_book = openpyxl.Workbook()
sheet = excel_book.active
sheet.title = "Dados Extração Cassol"

sheet["A1"] = "CEP"
sheet["B1"] = "LINK"
sheet["C1"] = "TRANSPORTADORA"
sheet["D1"] = "TEMPO"
sheet["E1"] = "CUSTO"

current_row = 2

driver.get(r'https://www.cassol.com.br/')

for i, row in tqdm(ceps.iterrows(), total=len(ceps), desc="Processando Cep"):

    num_cep = row['cep']
    print(num_cep)
    num_cep = str(num_cep)
    if driver:
        driver.quit()


    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(r'https://www.cassol.com.br/')

    """janelinha = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, '//div[contains(normalize-space(.), "Informe sua localização para acessar produtos e ofertas da sua região.")]')))
    """
    input_cep = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="modalInputCep"]')))
    input_cep.send_keys(str(num_cep))

    botao_cep = WebDriverWait(driver,10).until(EC.element_to_be_clickable(("xpath",'//button[contains(normalize-space(.),"Inserir CEP")]')))
    sleep(2)
    botao_cep.click()
    try:
        driver.get('https://www.cassol.com.br/piso-ceramico-acacia-madeira-retificado-67x67-ilhabella-/p)
        linha_transp = WebDriverWait(driver,5).until(EC.presence_of_element_located((By.XPATH,'//tr[contains(.,"Transportadora")]')))
    except:
        continue
    for i, row in tqdm(links.iterrows(), total=len(links), desc="Processando Links"):
        link = row['links']
        transportadora = "Transportadora"
        driver.get(link)
        tempo_ent = ''
        preco = ''
        
        try:
            linha_transp = WebDriverWait(driver,3).until(EC.presence_of_element_located((By.XPATH,'//tr[contains(.,"Transportadora")]')))
            
            linha_transp_html = linha_transp.get_attribute("outerHTML")
            linha_tratada = BeautifulSoup(linha_transp_html, "html.parser")
            tds_linha = linha_tratada.find_all("td")
            tempo = tds_linha[1].text
            preco = tds_linha[2].text

            print(preco)
            print(tempo)
            #print(tds_linha)
        except:
            tempo = 'não encontrado'
            preco = 'não encontrado'
            pass
        
        





        sheet.cell(row=current_row, column=1, value=(num_cep))
        sheet.cell(row=current_row, column=2, value=(link))
        sheet.cell(row=current_row, column=3, value=(transportadora))
        sheet.cell(row=current_row, column=4, value=(tempo))
        sheet.cell(row=current_row, column=5, value=(preco))

        current_row += 1

excel_book.save("teste_2.xlsx")


"""

sheet.cell(row=current_row, column=1, value=(nota))
sheet.cell(row=current_row, column=2, value="".join(awbfinal))
sheet.cell(row=current_row, column=3, value=" ; ".join(status_list))
sheet.cell(row=current_row, column=4, value=" ; ".join(data_historico))
sheet.cell(row=current_row, column=5, value=(tipo))
sheet.cell(row=current_row, column=6, value="".join(previsão_entrega))
sheet.cell(row=current_row, column=7, value="".join(previsão_entrega_atualizada))
sheet.cell(row=current_row, column=8, value="".join(ultimo_status))



cep = driver.find_element('xpath','//*[@id="modalInputCep"]')
cep.s
"""