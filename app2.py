from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from selenium.webdriver.firefox.options import Options
import pandas as pd
import openpyxl
from time import sleep
from tqdm import tqdm

# Forçar a leitura dos CEPs como strings
links = pd.read_excel(r"./planilha_teste.xlsx")
ceps = pd.read_excel(r"./planilha_nova.xlsx", sheet_name="ceps", dtype={'cep': str})

excel_book = openpyxl.Workbook()
sheet = excel_book.active
sheet.title = "Dados Extração Cassol"

sheet["A1"] = "CEP"
sheet["B1"] = "LINK"
sheet["C1"] = "TRANSPORTADORA"
sheet["D1"] = "TEMPO"
sheet["E1"] = "CUSTO"

current_row = 2

driver = webdriver.Chrome()
driver.get(r'https://www.cassol.com.br/')



try:
    linha_transp = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//tr[contains(.,"Entrega SC")]')))

    linha_transp_html = linha_transp.get_attribute("outerHTML")
    linha_tratada = BeautifulSoup(linha_transp_html, "html.parser")
    tds_linha = linha_tratada.find_all("td")
    tempo = tds_linha[1].text
    preco = tds_linha[2].text

    print(preco)
    print(tempo)
    transportadora = "Transportadora"
    
    sheet.cell(row=current_row, column=3, value=transportadora)
    sheet.cell(row=current_row, column=4, value=tempo)
    sheet.cell(row=current_row, column=5, value=preco)

    current_row += 1

except:
    pass